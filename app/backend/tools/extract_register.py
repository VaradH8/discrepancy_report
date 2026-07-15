"""
Regenerate equipment_registered.txt from the master equipment list.

RULE (since 2026-07-15): the register is the WHOLE EquipmentName column of the
latest "Equipment List to Inventive" export — every name counts, highlighted or
not. (Before this the rule was the yellow-highlighted subset only; the user
reversed that decision.)

Usage:
    python tools/extract_register.py "C:\\path\\to\\Equipment List to Inventive (4).csv"
    python tools/extract_register.py "C:\\path\\to\\Equipment List to Inventive (4).xlsx"

Writes backend/equipment_registered.txt (one EquipmentName per line, sorted, unique).
"""
import csv, os, sys

HEADER = "EquipmentName"
OUT = os.path.join(os.path.dirname(__file__), "..", "equipment_registered.txt")


def extract_csv(path):
    names = []
    with open(path, encoding="cp1252", errors="replace", newline="") as f:
        r = csv.reader(f)
        hdr = next(r)
        col = hdr.index(HEADER)
        for row in r:
            if len(row) > col:
                v = row[col].strip()
                if v:
                    names.append(v)
    return names


def extract_xlsx(path):
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    header = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    col = header.index(HEADER) + 1
    names = []
    for r in range(2, ws.max_row + 1):
        v = ws.cell(r, col).value
        v = str(v).strip() if v is not None else ""
        if v:
            names.append(v)
    return names


if __name__ == "__main__":
    if len(sys.argv) != 2:
        sys.exit(f"usage: python {sys.argv[0]} <Equipment List to Inventive .csv|.xlsx>")
    src = sys.argv[1]
    names = extract_csv(src) if src.lower().endswith(".csv") else extract_xlsx(src)
    uniq = sorted(set(names))
    out = os.path.normpath(OUT)
    with open(out, "w", encoding="utf-8") as f:
        f.write("\n".join(uniq) + "\n")
    print(f"column values: {len(names)}  unique written: {len(uniq)}  -> {out}")
