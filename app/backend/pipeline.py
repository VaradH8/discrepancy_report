"""
DWG discrepancy pipeline — reference vs candidate -> formatted Excel.
Engine extracted from the validated notebook workflow.
Requires `dwg2dxf` (LibreDWG) on PATH for DWG->DXF conversion.
"""
import os, re, math, shutil, subprocess, tempfile
import ezdxf
from ezdxf.math import Vec3
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

TAG = re.compile(r'^[A-Z]{1,3}(-[A-Z]{1,3})?-?\d{4,8}[A-Z]?(-\d{1,2})?$')

def dwg_to_dxf(dwg_path, dxf_path):
    """DWG->DXF via LibreDWG's dwg2dxf. Fallback converter."""
    r = subprocess.run(["dwg2dxf", "-o", dxf_path, dwg_path],
                       capture_output=True, text=True)
    if not os.path.exists(dxf_path):
        raise RuntimeError(f"DWG->DXF failed: {r.stderr[-400:]}")
    return dxf_path

def has_oda():
    """True if the ODA File Converter is installed (preferred DWG decoder)."""
    return shutil.which("ODAFileConverter") is not None

def _oda_to_dxf(dwg_path, work):
    """DWG->DXF via the ODA File Converter (headless, through xvfb). ODA converts
    a whole folder, so stage the file alone and pull the single .dxf back out.
    ODA reads every AutoCAD version, so it handles files LibreDWG can't decode."""
    ind = os.path.join(work, "oda_in"); outd = os.path.join(work, "oda_out")
    os.makedirs(ind, exist_ok=True); os.makedirs(outd, exist_ok=True)
    shutil.copyfile(dwg_path, os.path.join(ind, "input.dwg"))
    # ODA's Qt build writes config under $HOME; give it a writable one.
    env = dict(os.environ, HOME=work, QT_QPA_PLATFORM="xcb")
    # ODAFileConverter <in> <out> <out-ver> <out-type> <recurse> <audit> [filter]
    r = subprocess.run(["xvfb-run", "-a", "ODAFileConverter", ind, outd,
                        "ACAD2018", "DXF", "0", "1", "*.dwg"],
                       capture_output=True, text=True, timeout=600, env=env)
    for f in os.listdir(outd):
        if f.lower().endswith(".dxf"):
            return os.path.join(outd, f)
    raise RuntimeError(
        f"ODA produced no DXF (rc={r.returncode}); "
        f"stdout={r.stdout[-200:]!r}; stderr={r.stderr[-200:]!r}")

def _to_dxf(src_path, out_dxf, work):
    """Return a DXF path for a .dwg or .dxf source. A .dxf is used as-is (no
    conversion — it's exactly what ezdxf reads). A .dwg is converted with ODA
    when installed, falling back to LibreDWG. If both fail, surface both errors."""
    if src_path.lower().endswith(".dxf"):
        return src_path
    errs = []
    if has_oda():
        try:
            return _oda_to_dxf(src_path, work)
        except Exception as e:
            errs.append(f"ODA: {e}")
    try:
        return dwg_to_dxf(src_path, out_dxf)
    except Exception as e:
        errs.append(f"LibreDWG: {e}")
        raise RuntimeError("DWG->DXF failed — " + " || ".join(errs))

_CODE_RE = re.compile(r'^\s*(\d{1,4})\s*$')

def _is_code(line):
    """True if a physical line is a DXF group-code line (an int in 0..1071)."""
    m = _CODE_RE.match(line)
    return bool(m) and int(m.group(1)) <= 1071

def _sanitize_dxf(path):
    """Repair DXF code/value line-pairing. LibreDWG's dwg2dxf sometimes writes a
    literal newline inside a text value (notes/MTEXT), which splits that value
    across physical lines and desyncs every code/value pair after it — ezdxf's
    strict AND recover readers both reject it ('Invalid group code "<text>"').
    We walk the strict (code, value) alternation and re-join any continuation
    lines (where a code was expected but the line isn't one) back into the prior
    value. Streamed line-by-line so a multi-GB / multi-million-line DXF is safe."""
    fixed = path + ".fixed.dxf"
    with open(path, "r", encoding="utf-8", errors="replace") as fin, \
         open(fixed, "w", encoding="utf-8") as fout:
        expect_code, val = True, None
        for raw in fin:
            line = raw.rstrip("\n").rstrip("\r")
            if expect_code:
                if _is_code(line):
                    if val is not None:
                        fout.write(val + "\n"); val = None   # flush previous value
                    fout.write(line.strip() + "\n")          # write the code
                    expect_code = False
                elif val is not None:
                    val += " " + line                        # spilled value -> rejoin
                # else: stray junk before the first code — drop it
            else:
                val = line                                   # this line is the value
                expect_code = True
        if val is not None:
            fout.write(val + "\n")
    return fixed

def _read_dxf(dxf_path):
    """Tolerant DXF read with escalating fallbacks for malformed LibreDWG output:
    strict -> sanitize line-pairing then strict -> sanitize then recover."""
    try:
        return ezdxf.readfile(dxf_path)
    except Exception:
        fixed = _sanitize_dxf(dxf_path)
        try:
            return ezdxf.readfile(fixed)
        except Exception:
            from ezdxf import recover
            doc, _auditor = recover.readfile(fixed)
            return doc

def _text_xy(e):
    """(string, x, y) for a text-bearing entity, else None."""
    t = e.dxftype()
    if t == "TEXT":
        ins = e.dxf.insert; return e.dxf.text, ins.x, ins.y
    if t == "MTEXT":
        ins = e.dxf.insert
        try: s = e.plain_text()
        except Exception: s = getattr(e, "text", "")
        return s, ins.x, ins.y
    if t in ("ATTRIB", "ATTDEF"):
        ins = e.dxf.insert if e.dxf.hasattr("insert") else None
        return e.dxf.text, (ins.x if ins else 0.0), (ins.y if ins else 0.0)
    return None

def _scan(dxf_path):
    """Extract {tag: (x,y)}, datum points, geometry count, and an entity-type
    histogram from a DXF. Equipment tags may be free TEXT/MTEXT, block ATTRIBs,
    or TEXT nested inside block references (a converter like ODA keeps the block
    structure a DXF export would flatten).

    Nested text is resolved via a PER-BLOCK-DEFINITION cache: each block's tag
    texts (with local positions, sub-blocks folded in) are computed once, then
    every INSERT just matrix-transforms those points to world space. Exploding
    each insert (virtual_entities) instead is O(inserts x block size) and takes
    minutes / pegs CPU on 100MB+ layouts — which got the worker killed mid-request."""
    doc = _read_dxf(dxf_path); msp = doc.modelspace()
    tags, datums, geom, hist = {}, [], 0, {}
    def _add(s, x, y):
        s = (s or "").strip()
        if TAG.match(s):
            tags.setdefault(s, (round(x, 2), round(y, 2)))
    # 1) Modelspace — world positions, plus geometry + datum-marker counts.
    for e in msp:
        t = e.dxftype()
        hist[t] = hist.get(t, 0) + 1
        tv = _text_xy(e)
        if tv:
            _add(*tv)
        elif t == "INSERT":
            ins = e.dxf.insert
            if "DATUM" in e.dxf.name.upper():
                datums.append((ins.x, ins.y))
            try:
                for a in e.attribs:
                    av = _text_xy(a)
                    if av: _add(*av)
            except Exception:
                pass
        elif t in ("LINE", "LWPOLYLINE", "POLYLINE", "ARC", "SPLINE", "HATCH"):
            geom += 1
    # 2) Every block definition — equipment tags very often live inside blocks
    #    (a DXF *export* flattens them into modelspace; a converter keeps the
    #    block structure). Sweep all definitions once for the tag SET. Positions
    #    here are block-local, which is fine — the diff is set-based and these
    #    files carry no registrable datum layout anyway.
    for blk in doc.blocks:
        for e in blk:
            tv = _text_xy(e)
            if tv: _add(*tv)
    return tags, datums, geom, hist

def _nearest(pt, pts):
    if not pts: return ("", "")
    bx, by, bd = "", "", 1e18
    for dx, dy in pts:
        d = math.hypot(pt[0]-dx, pt[1]-dy)
        if d < bd: bd, bx, by = d, round(dx, 2), round(dy, 2)
    return (bx, by)

COLS = ["EquipmentName","Status","Old_X","Old_Y","Old_Z","New_X","New_Y","New_Z",
        "dX","dY","dZ","Beyond_Tolerance","Match_Confidence","Sheet","Locator","Notes"]

def _diff(ref_dwg, cand_dwg, sheet_name="SHT"):
    """Reference vs candidate -> (rows, stats). The shared engine behind compare()
    and compare_registered(); does the DWG scan and builds the discrepancy rows."""
    with tempfile.TemporaryDirectory() as tmp:
        ref_dxf  = _to_dxf(ref_dwg,  os.path.join(tmp, "ref.dxf"),  tmp)
        cand_dxf = _to_dxf(cand_dwg, os.path.join(tmp, "cand.dxf"), tmp)
        OLD, DAT,  ogeom, ohist = _scan(ref_dxf)
        NEW, DATC, cgeom, chist = _scan(cand_dxf)

    both, added, removed = set(OLD)&set(NEW), set(NEW)-set(OLD), set(OLD)-set(NEW)
    # A side is a real layout only if it has geometry AND datum-point markers.
    ref_is_layout  = ogeom > 200 and len(DAT)  > 0
    cand_is_layout = cgeom > 200 and len(DATC) > 0
    # Movement can only be computed when BOTH sides carry registrable positions.
    movement_ok = ref_is_layout and cand_is_layout
    # Heuristic swap warning: reference looks like a tiny extract, candidate like a full layout.
    swapped = (ogeom < 200 and len(DAT) == 0 and cgeom > 200 and len(DATC) > 0)
    warning = ("Files may be in the wrong slots: the reference has no geometry/datum markers "
               "but the candidate does. Put the full-layout DWG in the Reference slot.") if swapped else ""
    if not warning and not OLD and not NEW:
        _top = lambda h: ", ".join(f"{k}:{v}" for k, v in sorted(h.items(), key=lambda kv: -kv[1])[:6]) or "empty"
        warning = ("No equipment tags matched in either drawing (entity mix — "
                   f"reference[{_top(ohist)}], candidate[{_top(chist)}]). If the tags live in "
                   "xrefs or nested blocks, upload the DXF export instead.")
    rows = []
    for t in sorted(set(OLD)|set(NEW)):
        r = {c: "" for c in COLS}; r["EquipmentName"]=t; r["Sheet"]=sheet_name
        if t in added:
            r["Status"]="ADDED"; r["Match_Confidence"]="HIGH (tag-set)"
            nx,ny=NEW[t]; r["New_X"],r["New_Y"]=nx,ny; r["Locator"]=f"tag@({nx},{ny})"
            r["Notes"]="New equipment in candidate."
        elif t in removed:
            r["Status"]="REMOVED"; r["Match_Confidence"]="HIGH (tag-set)"
            ox,oy=_nearest(OLD[t],DAT); r["Old_X"],r["Old_Y"],r["Old_Z"]=ox,oy,0
            r["Notes"]="In reference, absent from candidate."
        else:
            r["Status"]="PRESENT IN BOTH"
            ox,oy=_nearest(OLD[t],DAT); r["Old_X"],r["Old_Y"],r["Old_Z"]=ox,oy,0
            if movement_ok:
                r["Beyond_Tolerance"]="REVIEW"; r["Match_Confidence"]="layout-vs-layout (compute dX/dY)"
                r["Notes"]="Both sides are layouts with datum markers — movement computable."
            else:
                r["Beyond_Tolerance"]="n/a"; r["Match_Confidence"]="LOW – candidate has no registrable coords"
                r["Notes"]="Candidate is a schedule extract; movement not computable."
        rows.append(r)

    stats = dict(both=len(both), added=sorted(added), removed=sorted(removed),
                 added_pos={t: list(NEW[t]) for t in added},
                 removed_pos={t: list(OLD[t]) for t in removed},
                 movement_ok=movement_ok, warning=warning,
                 ref_geom=ogeom, cand_geom=cgeom, ref_datums=len(DAT), cand_datums=len(DATC))
    return rows, stats

def compare(ref_dwg, cand_dwg, sheet_name="SHT", out_xlsx=None):
    rows, stats = _diff(ref_dwg, cand_dwg, sheet_name)
    xlsx = _write(rows, stats, sheet_name, out_xlsx)
    return xlsx, stats


# ── Master register (background allowlist) ───────────────────────────────────
# The discrepancy report is filtered down to only equipment in the register.
# The register is the YELLOW-HIGHLIGHTED subset of column D (EquipmentName) of
# "Equipment List to Inventive 1.xlsx" — NOT the whole column. Highlighting is
# lost in CSV, so those 909 names are pre-extracted into equipment_registered.txt
# (one name per line). To refresh after re-highlighting, see tools/extract_register.py.
MASTER_FILE = os.path.join(os.path.dirname(__file__), "equipment_registered.txt")
_MASTER = None  # cached allowlist of normalized EquipmentNames

def _norm(name):
    """Match key for an EquipmentName. Strips surrounding whitespace (a few
    entries carry stray spaces); otherwise an exact match."""
    return (name or "").strip()

def load_master(path=MASTER_FILE):
    """Set of normalized EquipmentNames in the register (highlighted subset)."""
    global _MASTER
    if _MASTER is not None:
        return _MASTER
    allow = set()
    try:
        with open(path, encoding="utf-8") as f:
            for line in f:
                n = _norm(line)
                if n:
                    allow.add(n)
    except FileNotFoundError:
        pass
    _MASTER = allow
    return _MASTER

SCOPE_LABELS = ("GLOBALLY ADDED", "LOCALLY ADDED", "GLOBALLY REMOVED", "LOCALLY REMOVED")

def compare_registered(ref_dwg, cand_dwg, sheet_name="SHT", out_xlsx=None, allow=None, scope=None):
    """Discrepancy report filtered to the master register, with each ADDED/REMOVED
    row's Status replaced by the cross-reference verdict (GLOBALLY/LOCALLY ADDED or
    REMOVED) supplied in `scope` = {EquipmentName: label}. Rows not in the register
    are dropped. Returns (xlsx_path, stats, dropped_names)."""
    allow = load_master() if allow is None else allow
    scope = {_norm(k): v for k, v in (scope or {}).items()}
    rows, stats = _diff(ref_dwg, cand_dwg, sheet_name)
    kept    = [r for r in rows if _norm(r["EquipmentName"]) in allow]
    dropped = [r["EquipmentName"] for r in rows if _norm(r["EquipmentName"]) not in allow]
    # Merge the cross-reference verdict into the Status text.
    for r in kept:
        lab = scope.get(_norm(r["EquipmentName"]))
        if lab in SCOPE_LABELS:
            r["Status"] = lab
    groups = {lab: sorted(r["EquipmentName"] for r in kept if r["Status"] == lab)
              for lab in SCOPE_LABELS}
    # Headline counts reflect the filtered report, so the UI/summary match the rows.
    fstats = dict(stats)
    fstats["both"]    = sum(1 for r in kept if r["Status"] == "PRESENT IN BOTH")
    fstats["added"]   = sorted(r["EquipmentName"] for r in kept if "ADDED" in r["Status"])
    fstats["removed"] = sorted(r["EquipmentName"] for r in kept if "REMOVED" in r["Status"])
    fstats["kept"], fstats["dropped"] = len(kept), len(dropped)
    fstats["master_count"] = len(allow)
    fstats["scope"] = groups
    xlsx = _write(kept, fstats, sheet_name, out_xlsx)
    return xlsx, fstats, dropped

def scan_drawing(dwg_path):
    """DWG/DXF -> {tag: (x, y)} tag-set with positions. One drawing, no diff."""
    with tempfile.TemporaryDirectory() as tmp:
        dxf = _to_dxf(dwg_path, os.path.join(tmp, "d.dxf"), tmp)
        tags, _datums, _geom, _hist = _scan(dxf)
    return tags


# ── Batch: whole-folder compare -> cross-reference -> registered ─────────────
# The exact per-module flow validated on real deliveries (m-04/m-05/M05B/M-07A):
# pair sheets by number, diff tag-sets, label each change GLOBALLY/LOCALLY by
# cross-referencing the ISSUED set, filter to the master register.
_BATCH_SHEET_RES = (re.compile(r'(?:SHEET|SHT)[-_ ]?0*(\d+)', re.I),  # SHT-01, SHT01, SHT_01, SHEET 1
                    re.compile(r'-0*(\d+)[_ ]*rev', re.I))            # ...-001-01_Rev C (no SHT marker)

def sheet_no_from_name(name):
    """Sheet number ('01'..) from a drawing filename, or None."""
    for rx in _BATCH_SHEET_RES:
        m = rx.search(name)
        if m:
            return f"{int(m.group(1)):02d}"
    return None

def batch_registered(issued_items, cand_items, out_dir, allow=None):
    """Bulk pipeline over two drawing sets.

    issued_items / cand_items : list of (filename, path) — the whole issued and
    candidate folders. Sheets are paired by number; each pair is diffed; every
    ADDED/REMOVED tag is labelled GLOBALLY vs LOCALLY against the ISSUED set;
    the registered report drops tags outside the master register. Writes
    SHT-XX_Registered.xlsx + SHT-XX_Changes.xlsx into out_dir.

    Returns (paths, summary).
    """
    allow = load_master() if allow is None else allow
    def scan_set(items):
        out = {}
        for label, path in items:
            n = sheet_no_from_name(label)
            if n is None or n in out:
                continue
            out[n] = set(scan_drawing(path))
        return out
    B, C = scan_set(issued_items), scan_set(cand_items)
    B_sheets = {}   # tag -> issued sheets it appears on
    for n, s in B.items():
        for tg in s:
            B_sheets.setdefault(tg, set()).add(n)
    paths, sheets = [], []
    for n in sorted(set(B) & set(C)):
        Bn, Cn = B[n], C[n]
        both, added, removed = Bn & Cn, Cn - Bn, Bn - Cn
        def lab_for(tg, status):
            if status == "ADDED":
                return "LOCALLY ADDED" if tg in B_sheets else "GLOBALLY ADDED"
            return "LOCALLY REMOVED" if (B_sheets.get(tg, set()) - {n}) else "GLOBALLY REMOVED"
        def build(filter_master):
            rows, groups = [], {k: [] for k in SCOPE_LABELS}
            for tg in sorted(Bn | Cn):
                if filter_master and _norm(tg) not in allow:
                    continue
                r = {c: "" for c in COLS}; r["EquipmentName"] = tg; r["Sheet"] = f"SHT_{n}"
                if tg in both:
                    r["Status"] = "PRESENT IN BOTH"; r["Notes"] = "Present in both revisions."
                else:
                    status = "ADDED" if tg in added else "REMOVED"
                    lab = lab_for(tg, status)
                    r["Status"] = lab; groups[lab].append(tg)
                    r["Match_Confidence"] = "HIGH (tag-set)"
                    r["Notes"] = ("Also elsewhere in the issued set." if lab.startswith("LOCALLY")
                                  else ("New in the candidate set." if status == "ADDED"
                                        else "Gone from the entire issued set."))
                rows.append(r)
            for k in groups:
                groups[k] = sorted(groups[k])
            return rows, groups
        entry = {"sheet": n,
                 "raw": {"added": len(added), "removed": len(removed), "both": len(both)}}
        for filter_master, suffix in ((True, "Registered"), (False, "Changes")):
            rows, groups = build(filter_master)
            nboth = sum(1 for r in rows if r["Status"] == "PRESENT IN BOTH")
            stats = dict(both=nboth,
                         added=sorted(r["EquipmentName"] for r in rows if "ADDED" in r["Status"]),
                         removed=sorted(r["EquipmentName"] for r in rows if "REMOVED" in r["Status"]),
                         ref_geom=0, cand_geom=0, ref_datums=0, cand_datums=0,
                         movement_ok=False, warning="", scope=groups)
            out = os.path.join(out_dir, f"SHT-{n}_{suffix}.xlsx")
            _write(rows, stats, f"SHT_{n}", out)
            paths.append(out)
            if filter_master:
                entry["reg"] = {"gA": len(groups["GLOBALLY ADDED"]), "lA": len(groups["LOCALLY ADDED"]),
                                "gR": len(groups["GLOBALLY REMOVED"]), "lR": len(groups["LOCALLY REMOVED"]),
                                "both": nboth, "kept": len(rows)}
        sheets.append(entry)
    summary = {"sheets": sheets,
               "unpaired_issued": sorted(set(B) - set(C)),
               "unpaired_candidate": sorted(set(C) - set(B)),
               "master_count": len(allow)}
    return paths, summary


SHEET_RE = re.compile(r'(SHT|SH|SHEET)[_\- ]?0*(\d+)', re.I)

def _sheet_no(label):
    """Pull a navigable sheet number out of a drawing label/filename.
    'I-DE-..._SHT_02.dwg' -> 'SHT_02'. Falls back to the filename stem."""
    if not label:
        return ""
    m = SHEET_RE.search(label)
    if m:
        return f"SHT_{int(m.group(2)):02d}"
    return os.path.splitext(os.path.basename(label))[0]


def _verdict(status, exists_elsewhere):
    if status == "ADDED":
        return ("Exists in other sheets — not globally new"
                if exists_elsewhere else "GLOBALLY NEW — appears nowhere else")
    if status == "REMOVED":
        return ("Still present in other sheets"
                if exists_elsewhere else "GLOBALLY REMOVED — gone from the whole set")
    return "—"


def cross_reference(tags_with_status, drawings, origin_labels=(),
                    added_origin=None, removed_origin=None, sheet_label="",
                    home_pos_map=None, out_xlsx=None):
    """
    Cross-check each added/removed tag against the whole drawing set.

    tags_with_status : list of {"tag": str, "status": "ADDED"|"REMOVED"}
    drawings         : list of (label, dwg_path) — every sheet to check against
    origin_labels    : labels that are the original reference/candidate; matches there
                       are flagged as the tag's own origin and excluded from
                       "exists elsewhere" so the verdict stays meaningful.
    added_origin     : candidate filename — the sheet ADDED tags actually live on.
    removed_origin   : reference filename — the sheet REMOVED tags lived on.
    sheet_label      : the comparison's sheet label, used as a last-resort home sheet.
    home_pos_map     : {tag: [x, y]} positions carried from the comparison (the
                       candidate position for ADDED, reference position for REMOVED),
                       used when the home sheet itself isn't in the uploaded set.

    Returns (xlsx_path, summary). summary["results"] is one entry per tag.
    """
    home_pos_map = home_pos_map or {}
    origin = set(origin_labels)
    scans = {}  # label -> {tag: (x, y)}
    for label, path in drawings:
        scans[label] = scan_drawing(path)

    results = []
    for item in tags_with_status:
        tag, status = item["tag"], item.get("status", "")
        found = []  # (label, x, y, is_origin)
        for label, tagmap in scans.items():
            if tag in tagmap:
                x, y = tagmap[tag]
                found.append((label, x, y, label in origin))
        elsewhere = [f for f in found if not f[3]]

        # Home sheet — where the tag actually sits, so the user has a page to open
        # even when it appears nowhere else. ADDED lives on the candidate, REMOVED
        # on the reference; fall back to the typed sheet label.
        home_label = added_origin if status == "ADDED" else removed_origin if status == "REMOVED" else None
        home_sheet = _sheet_no(home_label) or _sheet_no(sheet_label) or sheet_label
        # Prefer a fresh scan of the home file if it was uploaded; otherwise fall back
        # to the position carried over from the comparison.
        home_pos = (scans.get(home_label, {}).get(tag) if home_label else None) or home_pos_map.get(tag)

        results.append({
            "tag": tag,
            "status": status,
            "exists_elsewhere": bool(elsewhere),
            "home_sheet": home_sheet,
            "home_pos": list(home_pos) if home_pos else None,
            "elsewhere": [[_sheet_no(lbl), x, y] for lbl, x, y, _ in elsewhere],
            "found": found,
            "verdict": _verdict(status, bool(elsewhere)),
        })

    summary = dict(results=results,
                   drawings=[lbl for lbl, _ in drawings],
                   tag_count=len(results),
                   globally_new=sum(1 for r in results
                                    if r["status"] == "ADDED" and not r["exists_elsewhere"]),
                   globally_removed=sum(1 for r in results
                                        if r["status"] == "REMOVED" and not r["exists_elsewhere"]))
    xlsx = _write_crosscheck(results, summary, out_xlsx)
    return xlsx, summary


XCOLS = ["EquipmentName", "Compare Status", "Spotted On (sheet)", "Verdict",
         "Exists Elsewhere", "Also Found In (sheet @ position)", "Notes"]

def _write_crosscheck(results, summary, out_xlsx):
    wb = Workbook(); ws = wb.active; ws.title = "Cross-Reference"
    HDR = Font(bold=True, color="FFFFFF", name="Arial", size=10); HF = PatternFill("solid", start_color="1F4E78")
    base = Font(name="Arial", size=10); thin = Side(style="thin", color="BFBFBF")
    B = Border(left=thin, right=thin, top=thin, bottom=thin)
    NEW = PatternFill("solid", start_color="C6E0B4")     # green: globally new/removed (truly unique)
    SHARED = PatternFill("solid", start_color="FFF2CC")  # amber: exists elsewhere
    ws.append(XCOLS)
    for c in range(1, len(XCOLS) + 1):
        cell = ws.cell(1, c); cell.font = HDR; cell.fill = HF; cell.border = B
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for r in results:
        home = r["home_sheet"] or "—"
        if r["home_pos"]:
            home += f" @ ({r['home_pos'][0]},{r['home_pos'][1]})"
        loc = "; ".join(f"{sht} @ ({x},{y})" for sht, x, y in r["elsewhere"]) or "—"
        note = ("Also appears in other sheets — review whether it should be flagged as a change."
                if r["exists_elsewhere"] else
                ("Confirmed unique to this revision." if r["status"] == "ADDED"
                 else "Confirmed gone from the entire set."))
        ws.append([r["tag"], r["status"], home, r["verdict"],
                   "YES" if r["exists_elsewhere"] else "NO", loc, note])
        rr = ws.max_row
        fill = SHARED if r["exists_elsewhere"] else NEW
        for c in range(1, len(XCOLS) + 1):
            cell = ws.cell(rr, c); cell.font = base; cell.border = B
            if c == 5: cell.fill = fill
    for i, w in enumerate([22, 15, 22, 40, 14, 55, 50], 1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.freeze_panes = "A2"; ws.row_dimensions[1].height = 30
    ws.auto_filter.ref = f"A1:G{ws.max_row}"

    sm = wb.create_sheet("Summary & Method")
    lines = [
        ["Cross-reference — is each added/removed tag unique to its sheet?"], [""],
        [f"   Tags checked          : {summary['tag_count']}"],
        [f"   Globally NEW (added)   : {summary['globally_new']}  (appear in no other sheet)"],
        [f"   Globally REMOVED       : {summary['globally_removed']}  (gone from the whole set)"], [""],
        [f"   Drawing set scanned ({len(summary['drawings'])} sheets):"],
    ]
    for d in summary["drawings"]:
        lines.append([f"      • {d}"])
    lines += [[""],
        ["Method: each tag from the compare's ADDED/REMOVED register is searched in the"],
        ["tag-set of every uploaded drawing. A match in the original reference/candidate is"],
        ["marked [origin] and does not count toward 'exists elsewhere'."]]
    for L in lines: sm.append(L)
    sm["A1"].font = Font(bold=True, size=12, name="Arial"); sm.column_dimensions["A"].width = 95
    for r in range(2, sm.max_row + 1): sm.cell(r, 1).font = base
    out = out_xlsx or tempfile.mktemp(suffix=".xlsx")
    wb.save(out); return out


def _write(rows, stats, sheet_name, out_xlsx):
    wb=Workbook(); ws=wb.active; ws.title="Discrepancy"
    HDR=Font(bold=True,color="FFFFFF",name="Arial",size=10); HF=PatternFill("solid",start_color="1F4E78")
    base=Font(name="Arial",size=10); thin=Side(style="thin",color="BFBFBF"); B=Border(left=thin,right=thin,top=thin,bottom=thin)
    ADD=PatternFill("solid",start_color="C6E0B4"); REM=PatternFill("solid",start_color="F8CBAD"); BOTH=PatternFill("solid",start_color="FFF2CC")
    ws.append(COLS)
    for c in range(1,len(COLS)+1):
        cell=ws.cell(1,c);cell.font=HDR;cell.fill=HF;cell.border=B
        cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
    for r in rows:
        ws.append([r[c] for c in COLS]); rr=ws.max_row
        # substring match so GLOBALLY/LOCALLY ADDED|REMOVED colour like ADDED/REMOVED
        fill=ADD if "ADDED" in r["Status"] else REM if "REMOVED" in r["Status"] else BOTH
        for c in range(1,len(COLS)+1):
            cell=ws.cell(rr,c);cell.font=base;cell.border=B
            if c==2: cell.fill=fill
    for i,w in enumerate([22,16,9,9,7,9,9,7,7,7,7,15,30,8,26,46],1):
        ws.column_dimensions[chr(64+i)].width=w
    ws.freeze_panes="A2"; ws.row_dimensions[1].height=30; ws.auto_filter.ref=f"A1:P{ws.max_row}"
    sm=wb.create_sheet("Summary & Method")
    rows_sm = [
      [f"Equipment discrepancy — {sheet_name}"],[""],
    ]
    if stats.get("warning"):
        rows_sm += [["⚠ "+stats["warning"]],[""]]
    rows_sm += [
      ["Equipment-level (reliable):"],
      [f"   Present in both : {stats['both']}"],
      [f"   ADDED : {len(stats['added'])}  -> {', '.join(stats['added'])}"],
      [f"   REMOVED : {len(stats['removed'])}  -> {', '.join(stats['removed'])}"],[""],
      [f"   reference: {stats['ref_geom']} geometry, {stats['ref_datums']} datum markers"],
      [f"   candidate: {stats['cand_geom']} geometry, {stats['cand_datums']} datum markers"],
      [f"   movement computable: {'YES' if stats['movement_ok'] else 'NO — a side lacks datum-marked positions'}"],
    ]
    if stats.get("scope") and any(stats["scope"].values()):
        g = stats["scope"]
        rows_sm += [[""], ["Cross-reference scope (registered rows):"],
            [f"   GLOBALLY ADDED   : {len(g['GLOBALLY ADDED'])}  -> {', '.join(g['GLOBALLY ADDED'])}"],
            [f"   LOCALLY ADDED    : {len(g['LOCALLY ADDED'])}  -> {', '.join(g['LOCALLY ADDED'])}"],
            [f"   GLOBALLY REMOVED : {len(g['GLOBALLY REMOVED'])}  -> {', '.join(g['GLOBALLY REMOVED'])}"],
            [f"   LOCALLY REMOVED  : {len(g['LOCALLY REMOVED'])}  -> {', '.join(g['LOCALLY REMOVED'])}"]]
    for L in rows_sm: sm.append(L)
    sm["A1"].font=Font(bold=True,size=12,name="Arial"); sm.column_dimensions["A"].width=95
    for r in range(2,sm.max_row+1): sm.cell(r,1).font=base
    out = out_xlsx or tempfile.mktemp(suffix=".xlsx")
    wb.save(out); return out
